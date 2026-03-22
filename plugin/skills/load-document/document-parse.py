#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pymupdf",
#   "python-docx",
#   "openpyxl",
#   "path",
# ]
# ///
"""Parse contract documents (PDF, DOCX, XLSX) or ZIP archives into structured JSON clauses.

Usage:
    uv run document-parse.py <path> [--output <json-path>]

Supports:
    .pdf   — PyMuPDF structured clause extraction
    .docx  — python-docx paragraph/heading extraction
    .xlsx  — openpyxl spreadsheet-to-text extraction
    .zip   — extract and process all supported files within
    .doc   — skipped with warning (needs antiword/libreoffice)
    .xls   — skipped with warning (needs xlrd)

Outputs a JSON array of clause objects to stdout (or file if --output given).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import zipfile

import fitz  # PyMuPDF


# Patterns for line classification (PDF parsing)
# Section heading: "4. INTELLECTUAL PROPERTY" (number, dot, space, ALL CAPS title)
SECTION_RE = re.compile(r"^(\d+)\.\s+([A-Z][A-Z &,\-/]+)$")
# Clause number alone on a line: "2.1" or "12.10"
CLAUSE_BARE_RE = re.compile(r"^(\d+\.\d+)$")
# Clause number with inline text: "12.10 Force Majeure. Neither Party..."
CLAUSE_INLINE_RE = re.compile(r"^(\d+\.\d+)\s+(.+)")
# Exhibit sub-heading with text: "A.1 Overview"
SUB_INLINE_RE = re.compile(r"^([A-Z]\.\d+)\s+(.+)")
# Exhibit sub-heading bare: "A.1"
SUB_BARE_RE = re.compile(r"^([A-Z]\.\d+)$")
# Exhibit heading: "EXHIBIT A: Statement of Work #1"
EXHIBIT_RE = re.compile(r"^EXHIBIT\s+([A-Z])[.:]\s*(.*)", re.IGNORECASE)
# Preamble heading
PREAMBLE_RE = re.compile(r"^PREAMBLE$")
# TOC heading
TOC_RE = re.compile(r"^TABLE OF CONTENTS$")
# Signature block
SIGNATURE_RE = re.compile(r"^IN WITNESS WHEREOF", re.IGNORECASE)
# Page header/footer
HEADER_RE = re.compile(r"^CONFIDENTIAL\s*[-—–]")
FOOTER_RE = re.compile(r"^Page \d+/\d+$")

# Heading-like patterns for DOCX/generic parsing
HEADING_NUM_RE = re.compile(r"^(\d+(?:\.\d+)*)\s+(.*)")

# Heuristic keyword → flag mapping
FLAG_KEYWORDS: dict[str, list[str]] = {
    "ip": ["intellectual property", "work product", "assignment", "license",
           "copyright", "patent", "trademark", "moral rights", "pre-existing ip"],
    "termination": ["termination", "terminate", "term of", "expiration",
                    "renewal", "auto-renew"],
    "payment": ["fee", "payment", "invoice", "compensation", "rate",
                "reimburse", "expense"],
    "confidentiality": ["confidential", "non-disclosure", "trade secret",
                        "proprietary"],
    "liability": ["liability", "indemnif", "damages", "limitation of",
                  "consequential"],
    "staffing": ["staffing", "personnel", "team", "headcount", "resource"],
    "compliance": ["compliance", "regulatory", "gdpr", "ccpa", "privacy",
                   "data protection", "audit"],
    "noncompete": ["non-compet", "non-solicitation", "restrictive covenant"],
    "insurance": ["insurance", "coverage", "policy"],
    "governance": ["governance", "escalation", "change management",
                   "change control"],
    # RFP-relevant flags
    "scope": ["scope of work", "scope of service", "project scope",
              "work scope", "scope"],
    "requirements": ["requirement", "shall provide", "must provide",
                     "shall include", "minimum qualif", "mandatory"],
    "evaluation": ["evaluation", "scoring", "selection criteria",
                   "evaluation criteria", "weighted", "points"],
    "timeline": ["timeline", "schedule", "milestone", "deadline",
                 "due date", "delivery date", "completion date"],
    "pricing": ["pricing", "cost proposal", "fee schedule", "rate schedule",
                "price", "budget", "compensation"],
    "qualifications": ["qualification", "experience", "certified",
                       "certification", "license", "background"],
    "technical": ["technical", "specification", "architecture",
                  "infrastructure", "system", "technology"],
    "management": ["management", "project manager", "oversight",
                   "supervision", "reporting"],
    "submission": ["submission", "proposal due", "submit", "response format",
                   "proposal format", "submittal"],
}


def assign_flags(text: str) -> list[str]:
    """Assign topic flags based on keyword matching."""
    lower = text.lower()
    return sorted(
        flag for flag, keywords in FLAG_KEYWORDS.items()
        if any(kw in lower for kw in keywords)
    )


def _flush(current: dict | None, clauses: list[dict]):
    """Append current clause to list if it exists."""
    if current is not None:
        current["body"] = current["body"].strip()
        clauses.append(current)


def _new_clause(section_number: str, section_title: str, body: str,
                page: int, source_file: str = "") -> dict:
    return {
        "section_number": section_number,
        "section_title": section_title,
        "body": body,
        "source_file": source_file,
        "page_start": page,
        "page_end": page,
    }


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------

def extract_clauses_pdf(pdf_path: str, source_file: str = "") -> list[dict]:
    """Extract structured clauses from a contract PDF."""
    if not source_file:
        from path import Path as PathJ
        source_file = PathJ(pdf_path).name

    doc = fitz.open(pdf_path)

    # Collect all lines with their page numbers
    lines: list[tuple[int, str]] = []  # (page_1indexed, text)
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        for line in text.split("\n"):
            stripped = line.strip()
            if stripped:
                lines.append((page_num + 1, stripped))
    doc.close()

    clauses: list[dict] = []
    current: dict | None = None
    in_toc = False  # skip TOC content

    i = 0
    while i < len(lines):
        page, line = lines[i]

        # Skip headers and footers
        if HEADER_RE.match(line) or FOOTER_RE.match(line):
            i += 1
            continue

        # Detect and skip TOC
        if TOC_RE.match(line):
            in_toc = True
            i += 1
            continue

        # End TOC when we hit PREAMBLE or a section heading
        if in_toc:
            if PREAMBLE_RE.match(line) or SECTION_RE.match(line):
                in_toc = False
                # fall through to handle this line
            else:
                i += 1
                continue

        # PREAMBLE
        if PREAMBLE_RE.match(line):
            _flush(current, clauses)
            current = _new_clause("Preamble", "Preamble", "", page, source_file)
            i += 1
            continue

        # Signature block
        if SIGNATURE_RE.match(line):
            _flush(current, clauses)
            current = _new_clause("Signatures", "Signatures", "", page, source_file)
            i += 1
            continue

        # Section heading: "4. INTELLECTUAL PROPERTY"
        m = SECTION_RE.match(line)
        if m:
            _flush(current, clauses)
            current = _new_clause(m.group(1), m.group(2).strip().title(),
                                  "", page, source_file)
            i += 1
            continue

        # Exhibit heading: "EXHIBIT A: ..."
        m = EXHIBIT_RE.match(line)
        if m:
            _flush(current, clauses)
            title = m.group(2).strip()
            current = _new_clause(f"Exhibit {m.group(1)}",
                                  title.title() if title else "", "", page,
                                  source_file)
            i += 1
            continue

        # Clause number bare on its own line: "2.1"
        m = CLAUSE_BARE_RE.match(line)
        if m:
            _flush(current, clauses)
            current = _new_clause(m.group(1), "", "", page, source_file)
            i += 1
            continue

        # Clause number with inline text: "12.10 Force Majeure. ..."
        m = CLAUSE_INLINE_RE.match(line)
        if m:
            _flush(current, clauses)
            # Try to extract title from "Title. Body text..."
            rest = m.group(2)
            title = ""
            dot_match = re.match(r"^([A-Z][^.]+)\.\s*(.*)", rest)
            if dot_match:
                title = dot_match.group(1)
                rest = dot_match.group(2)
            current = _new_clause(m.group(1), title, rest, page, source_file)
            i += 1
            continue

        # Exhibit sub-heading bare: "A.1"
        m = SUB_BARE_RE.match(line)
        if m:
            _flush(current, clauses)
            current = _new_clause(m.group(1), "", "", page, source_file)
            i += 1
            continue

        # Exhibit sub-heading with text: "A.1 Overview"
        m = SUB_INLINE_RE.match(line)
        if m:
            _flush(current, clauses)
            current = _new_clause(m.group(1), m.group(2), "", page, source_file)
            i += 1
            continue

        # Continuation text — append to current clause body
        if current is not None:
            if current["body"]:
                current["body"] += " " + line
            else:
                current["body"] = line
            current["page_end"] = page

        i += 1

    _flush(current, clauses)

    # Post-process: assign flags
    for clause in clauses:
        combined = f"{clause.get('section_title', '')} {clause['body']}"
        clause["flags"] = assign_flags(combined)

    return clauses


# ---------------------------------------------------------------------------
# DOCX parsing
# ---------------------------------------------------------------------------

def extract_clauses_docx(docx_path: str, source_file: str = "") -> list[dict]:
    """Extract structured sections from a DOCX file."""
    from docx import Document

    if not source_file:
        from path import Path as PathJ
        source_file = PathJ(docx_path).name

    doc = Document(docx_path)
    clauses: list[dict] = []
    current: dict | None = None
    section_counter = 0

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        is_heading = (para.style and para.style.name
                      and para.style.name.lower().startswith("heading"))

        # Check for numbered headings in text itself
        num_match = HEADING_NUM_RE.match(text)

        if is_heading or num_match:
            _flush(current, clauses)
            if num_match:
                sec_num = num_match.group(1)
                sec_title = num_match.group(2).strip()
            else:
                section_counter += 1
                sec_num = str(section_counter)
                sec_title = text
            current = _new_clause(sec_num, sec_title, "", 1, source_file)
        elif current is not None:
            if current["body"]:
                current["body"] += "\n" + text
            else:
                current["body"] = text
        else:
            # No heading seen yet — create an intro section
            section_counter += 1
            current = _new_clause(str(section_counter), "", text, 1, source_file)

    _flush(current, clauses)

    # Assign flags
    for clause in clauses:
        combined = f"{clause.get('section_title', '')} {clause['body']}"
        clause["flags"] = assign_flags(combined)

    return clauses


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

def extract_clauses_xlsx(xlsx_path: str, source_file: str = "") -> list[dict]:
    """Extract text content from an XLSX spreadsheet, one section per sheet."""
    from openpyxl import load_workbook

    if not source_file:
        from path import Path as PathJ
        source_file = PathJ(xlsx_path).name

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  Warning: could not open {source_file}: {exc}", file=sys.stderr)
        return []

    clauses: list[dict] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(c for c in cells if c)
            if line.strip():
                rows_text.append(line)

        if not rows_text:
            continue

        body = "\n".join(rows_text)
        clause = _new_clause(
            section_number=f"Sheet-{sheet_idx}",
            section_title=sheet_name,
            body=body,
            page=1,
            source_file=source_file,
        )
        combined = f"{sheet_name} {body}"
        clause["flags"] = assign_flags(combined)
        clauses.append(clause)

    wb.close()
    return clauses


# ---------------------------------------------------------------------------
# ZIP extraction and dispatch
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
SKIP_WARN_EXTENSIONS = {".doc", ".xls"}


def extract_clauses_from_zip(zip_path: str) -> list[dict]:
    """Extract a ZIP archive and parse each supported file inside."""
    from path import Path as PathJ

    all_clauses: list[dict] = []
    skipped: list[str] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(tmpdir)

        # Walk extracted files
        tmp_path = PathJ(tmpdir)
        files = sorted(
            f for f in tmp_path.walkfiles()
            if not f.name.startswith(".")
            and not f.name.startswith("__")
        )

        for fpath in files:
            ext = fpath.suffix.lower()
            fname = fpath.name

            if ext == ".pdf":
                print(f"  Parsing PDF: {fname}", file=sys.stderr)
                clauses = extract_clauses_pdf(str(fpath), source_file=fname)
                all_clauses.extend(clauses)
            elif ext == ".docx":
                print(f"  Parsing DOCX: {fname}", file=sys.stderr)
                clauses = extract_clauses_docx(str(fpath), source_file=fname)
                all_clauses.extend(clauses)
            elif ext == ".xlsx":
                print(f"  Parsing XLSX: {fname}", file=sys.stderr)
                clauses = extract_clauses_xlsx(str(fpath), source_file=fname)
                all_clauses.extend(clauses)
            elif ext == ".xls":
                print(f"  Warning: skipping legacy XLS file: {fname} "
                      "(pre-convert to XLSX for full support)", file=sys.stderr)
                skipped.append(fname)
            elif ext == ".doc":
                print(f"  Warning: skipping legacy DOC file: {fname} "
                      "(pre-convert to DOCX for full support)", file=sys.stderr)
                skipped.append(fname)
            else:
                print(f"  Skipping unsupported file: {fname}", file=sys.stderr)
                skipped.append(fname)

    if skipped:
        print(f"\n  Skipped {len(skipped)} unsupported files: "
              f"{', '.join(skipped)}", file=sys.stderr)

    return all_clauses


# ---------------------------------------------------------------------------
# Dispatch by extension
# ---------------------------------------------------------------------------

def extract_clauses(input_path: str) -> list[dict]:
    """Route to the appropriate parser based on file extension."""
    from path import Path as PathJ
    p = PathJ(input_path)
    ext = p.suffix.lower()

    if ext == ".zip":
        return extract_clauses_from_zip(input_path)
    elif ext == ".pdf":
        return extract_clauses_pdf(input_path)
    elif ext == ".docx":
        return extract_clauses_docx(input_path)
    elif ext == ".xlsx":
        return extract_clauses_xlsx(input_path)
    elif ext == ".xls":
        print(f"Warning: legacy XLS format — pre-convert to XLSX for full support",
              file=sys.stderr)
        return []
    elif ext == ".doc":
        print(f"Warning: legacy DOC format — pre-convert to DOCX for full support",
              file=sys.stderr)
        return []
    else:
        print(f"Error: unsupported file format: {ext}", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Parse contract documents (PDF/DOCX/XLSX/ZIP) to JSON")
    parser.add_argument("input_path",
                        help="Path to document file or ZIP archive")
    parser.add_argument("--output", "-o",
                        help="Output JSON file (default: stdout)")
    args = parser.parse_args()

    clauses = extract_clauses(args.input_path)
    result = json.dumps(clauses, indent=2, ensure_ascii=False)

    if args.output:
        from path import Path as PathJ
        PathJ(args.output).parent.makedirs_p()
        PathJ(args.output).write_text(result)
        print(f"Wrote {len(clauses)} clauses to {args.output}", file=sys.stderr)
    else:
        print(result)

    # Summary to stderr
    pages = set()
    source_files = set()
    for c in clauses:
        pages.update(range(c["page_start"], c["page_end"] + 1))
        if c.get("source_file"):
            source_files.add(c["source_file"])

    print(f"\nSummary: {len(clauses)} clauses across {len(pages)} pages "
          f"from {len(source_files)} documents", file=sys.stderr)
    if source_files:
        for sf in sorted(source_files):
            count = sum(1 for c in clauses if c.get("source_file") == sf)
            print(f"  {sf}: {count} clauses", file=sys.stderr)


if __name__ == "__main__":
    main()
